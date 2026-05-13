"""XLSX export — openpyxl."""

from __future__ import annotations

from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill


def build_xlsx(path: Path, title: str, sections: list[dict]) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Summary"
    ws["A1"] = title
    ws["A1"].font = Font(bold=True, size=14)
    ws.merge_cells("A1:F1")
    ws["A1"].alignment = Alignment(horizontal="center")

    row = 3
    for sec in sections:
        ws.cell(row=row, column=1, value=sec.get("heading", "")).font = Font(bold=True, size=12)
        row += 1
        if sec.get("markdown"):
            for line in sec["markdown"].splitlines():
                ws.cell(row=row, column=1, value=line)
                row += 1
            row += 1
        if sec.get("table"):
            rows = sec["table"]
            if rows:
                cols = list(rows[0].keys())
                for i, c in enumerate(cols, start=1):
                    cell = ws.cell(row=row, column=i, value=c)
                    cell.font = Font(bold=True)
                    cell.fill = PatternFill("solid", fgColor="DDDDDD")
                row += 1
                for r in rows:
                    for i, c in enumerate(cols, start=1):
                        ws.cell(row=row, column=i, value=r.get(c, ""))
                    row += 1
                row += 1

    # OHLC sheet from any candle chart found
    for sec in sections:
        spec = sec.get("chart_spec")
        if spec and spec.get("kind") == "candle" and spec.get("bars"):
            sym = spec.get("symbol") or "Bars"
            sheet = wb.create_sheet(sym[:30])
            headers = ["timestamp", "open", "high", "low", "close", "volume"]
            for i, h in enumerate(headers, start=1):
                cell = sheet.cell(row=1, column=i, value=h)
                cell.font = Font(bold=True)
            for r_i, bar in enumerate(spec["bars"], start=2):
                for c_i, h in enumerate(headers, start=1):
                    sheet.cell(row=r_i, column=c_i, value=bar.get(h))

    # Auto-ish widths
    for sheet in wb.worksheets:
        for col_cells in sheet.columns:
            try:
                max_len = max(len(str(c.value)) if c.value is not None else 0 for c in col_cells)
                sheet.column_dimensions[col_cells[0].column_letter].width = min(max_len + 2, 40)
            except Exception:
                continue
    wb.save(path)
